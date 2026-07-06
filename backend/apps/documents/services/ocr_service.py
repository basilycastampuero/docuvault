import io
import logging
from typing import TYPE_CHECKING

import pytesseract
from botocore.exceptions import ClientError
from django.conf import settings
from pdf2image import convert_from_bytes
from PIL import Image

from apps.audit.models import AuditAction
from apps.audit.services import audit_service
from apps.core.exceptions import TransientError
from apps.documents.models import OcrStatus
from apps.documents.storage import StorageService

if TYPE_CHECKING:
    from apps.documents.models import Document

logger = logging.getLogger(__name__)

_IMAGE_MIMES = frozenset({"image/jpeg", "image/png"})
_PDF_MIME = "application/pdf"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_OFFICE_MIMES = frozenset({_DOCX_MIME, _XLSX_MIME})

# boto3 error codes that mean the blob is simply gone — retrying cannot help.
_PERMANENT_STORAGE_CODES = frozenset({"NoSuchKey", "404", "NoSuchBucket"})


def process(document: "Document") -> None:
    """Extract text from a document via OCR/parsing and make it searchable.

    Idempotent: safe to run more than once (overwrites the previous result), since
    Celery may re-deliver a message. PDFs and images are OCR'd via Tesseract; DOCX/XLSX
    (OOXML only, not legacy .doc/.xls) have their text extracted directly; everything
    else is marked skipped. Saving ocr_content triggers the FTS signal (Phase 3.3), so
    the document becomes searchable by its content with no extra indexing code.

    Raises TransientError on recoverable failures (e.g. a storage timeout) so the
    task retries; permanent failures (missing blob, corrupt file) set ocr_status to
    failed and return without retrying.
    """
    _set_status(document, OcrStatus.PROCESSING)

    is_supported = (
        document.mime_type in _IMAGE_MIMES
        or document.mime_type == _PDF_MIME
        or document.mime_type in _OFFICE_MIMES
    )
    if not is_supported:
        _set_status(document, OcrStatus.SKIPPED)
        logger.info("OCR skipped for %s (mime=%s)", document.pk, document.mime_type)
        return

    blob = _download(document)
    if blob is None:
        return  # already marked failed by _download

    try:
        text = _extract_text(document.mime_type, blob)
    except Exception:
        _set_status(document, OcrStatus.FAILED)
        logger.exception("OCR extraction failed for %s", document.pk)
        return

    document.ocr_content = text
    document.ocr_status = OcrStatus.COMPLETED
    document.save(update_fields=["ocr_content", "ocr_status", "updated_at"])

    audit_service.log(
        organization=document.organization,
        user=None,
        entity_type="document",
        entity_id=str(document.pk),
        action=AuditAction.UPDATE,
        new_values={"ocr_status": OcrStatus.COMPLETED.value},
        metadata={"via": "ocr"},
    )
    logger.info("OCR completed for %s (%d chars)", document.pk, len(text))


def _download(document: "Document") -> bytes | None:
    """Fetch the blob. Returns None (and marks failed) if it is permanently gone."""
    storage = StorageService()
    try:
        return storage.download_file(document.storage_path)
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code")
        if code in _PERMANENT_STORAGE_CODES:
            _set_status(document, OcrStatus.FAILED)
            logger.warning(
                "OCR blob missing for %s (%s); marking failed", document.pk, code
            )
            return None
        raise TransientError(
            f"Storage error downloading {document.pk}: {code}"
        ) from exc
    except Exception as exc:  # network/timeout — recoverable
        raise TransientError(f"Storage unavailable for {document.pk}") from exc


def _extract_text(mime_type: str, blob: bytes) -> str:
    """Dispatch to the right extractor (OCR for images/PDF, direct parsing for Office)."""
    if mime_type in _IMAGE_MIMES:
        image = Image.open(io.BytesIO(blob))
        return pytesseract.image_to_string(image, lang=settings.OCR_LANGUAGES).strip()

    if mime_type == _PDF_MIME:
        pages = convert_from_bytes(blob, dpi=settings.OCR_PDF_DPI)
        texts = [
            pytesseract.image_to_string(page, lang=settings.OCR_LANGUAGES)
            for page in pages
        ]
        return "\n".join(texts).strip()

    if mime_type == _DOCX_MIME:
        return _extract_docx(blob)

    return _extract_xlsx(blob)


def _extract_docx(blob: bytes) -> str:
    """Concatenate the text of every paragraph in a DOCX file."""
    from docx import Document as DocxDocument

    doc = DocxDocument(io.BytesIO(blob))
    return "\n".join(p.text for p in doc.paragraphs).strip()


def _extract_xlsx(blob: bytes) -> str:
    """Concatenate every cell value across all sheets of an XLSX file."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        lines = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(value) for value in row if value is not None]
                if cells:
                    lines.append("\t".join(cells))
        return "\n".join(lines).strip()
    finally:
        workbook.close()


def _set_status(document: "Document", ocr_status: OcrStatus) -> None:
    """Persist a status transition without touching searchable fields."""
    document.ocr_status = ocr_status
    document.save(update_fields=["ocr_status", "updated_at"])
